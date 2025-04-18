# 엑셀 저장 함수 
import openpyxl
from openpyxl.styles import Alignment
from openpyxl import Workbook

def save_to_excel(data: list[dict], path: str):
    

    wb = Workbook()
    ws = wb.active
    ws.title = "ScanData"

    # 헤더 작성
    if not data:
        return
    headers = list(data[0].keys())
    ws.append(headers)

    # 데이터 작성
    for row in data:
        ws.append([row.get(h, "") for h in headers])

    wb.save(path)
    print(f"[Excel Saved] {path}")

    # wb = openpyxl.Workbook()
    # ws = wb.active
    # ws.title = "Scan List"

    # # 헤더
    # headers = ["Check", "Roll", "Seq", "Shot", "Version", "Type", "Path", "Scan", "Clip"]
    # ws.append(headers)

    # # 데이터 쓰기
    # for d in data_list:
    #     ws.append([
    #         "Y" if d.get("check") else "N",
    #         d.get("roll", ""),
    #         d.get("seq", ""),
    #         d.get("shot", ""),
    #         d.get("version", ""),
    #         d.get("type", ""),
    #         d.get("path", ""),
    #         d.get("scan", ""),
    #         d.get("clip", "")
    #     ])

    # # 스타일
    # for row in ws.iter_rows(min_row=2):
    #     for cell in row:
    #         cell.alignment = Alignment(horizontal="center")

    # wb.save(path)
    
def load_shotnames_from_excel(path: str) -> dict:
    # exr 경로 : 샷 이름
    mapping = {}
    wb = openpyxl.load_workbook(path)
    ws = wb.active

    for row in ws.iter_rows(min_row=2):
        exr_path = row[0].value
        shot_name = row[1].value
        if exr_path and shot_name:
            mapping[exr_path] = shot_name

    return mapping
